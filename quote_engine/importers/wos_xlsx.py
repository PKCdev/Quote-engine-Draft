from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class SheetItem:
    material: str
    qty: int
    thickness: Optional[str] = None
    sheet_size: Optional[str] = None


@dataclass
class BandItem:
    spec: str
    lm: float


@dataclass
class HardwareItem:
    description: str
    qty: int


def _cells(ws) -> List[List[str]]:
    rows: List[List[str]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in r])
    return rows


def _cell_texts(row: List[str]) -> List[str]:
    return [c for c in row if c]


def parse(path: Optional[Path]) -> Dict[str, Any]:
    if not path:
        return {"sheets": [], "edgeband": [], "hardware": []}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = _cells(ws)

    # Locate anchors (collect all hardware anchors)
    anchors: Dict[str, int] = {}
    hw_anchors: List[int] = []
    for idx, row in enumerate(rows):
        text = " ".join(_cell_texts(row)).lower()
        if ("sheet stock totals" in text) and ("optimized" in text):
            anchors["sheet"] = idx
        if "edgeband totals" in text:
            anchors["edgeband"] = idx
        if "hardware totals" in text:
            hw_anchors.append(idx)

    sheet_items: List[SheetItem] = []
    band_items: List[BandItem] = []
    hw_items: List[HardwareItem] = []

    # Parse sheet stock: rows typically contain 'Qty - N', 'Thick - X', 'Sheet Size - W x H', and material name in another cell.
    if "sheet" in anchors:
        i = anchors["sheet"] + 1
        import re
        while i < len(rows):
            row = rows[i]
            cells = _cell_texts(row)
            text = " ".join(cells)
            ltext = text.lower()
            if not text:
                i += 1
                continue
            if "buyout totals" in ltext or "edgeband totals" in ltext:
                break
            if "qty -" in ltext:
                try:
                    qty = int(ltext.split("qty -", 1)[1].split()[0].replace(",", ""))
                    # Extract thickness and sheet size if present
                    thickness = None
                    sheet_size = None
                    material = None
                    for c in cells:
                        lc = c.lower()
                        if lc.startswith("thick -"):
                            try:
                                thickness = lc.split("thick -", 1)[1].strip().split()[0]
                            except Exception:
                                pass
                        elif lc.startswith("sheet size -"):
                            sheet_size = c.split("-", 1)[1].strip()
                        elif not any(k in lc for k in ["qty -", "thick -", "sheet size -", "total qty -", "work order", "project name:"]):
                            # heuristic: the longest remaining cell is material name
                            material = c if (material is None or len(c) > len(material)) else material
                    material = material or "Unknown material"
                    sheet_items.append(SheetItem(material=material, qty=qty, thickness=thickness, sheet_size=sheet_size))
                except Exception:
                    pass
            i += 1

    # Parse edgeband totals: best-effort
    if "edgeband" in anchors:
        i = anchors["edgeband"] + 1
        pending_specs: List[str] = []
        while i < len(rows):
            row = rows[i]
            cells = _cell_texts(row)
            text = " ".join(cells)
            ltext = text.lower()
            if not text:
                i += 1
                continue
            if "hardware totals" in ltext:
                break
            # Band spec row often contains width/thickness and a name in mid columns
            if any(tok in ltext for tok in [" 1mm", "1 mm", "x1mm", "x 1mm"]):
                # try to pick the cell following the thickness indicator, else longest candidate
                spec = None
                try:
                    idx_1mm = next(i for i,c in enumerate(cells) if "1mm" in c.lower())
                    if idx_1mm + 1 < len(cells):
                        spec = cells[idx_1mm + 1].strip()
                except StopIteration:
                    pass
                if not spec:
                    spec_candidates = [c for c in cells if len(c) > 3 and "mm" in c.lower()]
                    spec = max(spec_candidates, key=len, default="").strip()
                if spec:
                    pending_specs.append(spec)
            if ("meters -" in ltext) or ("metres -" in ltext):
                try:
                    m = re.search(r"(?:met(?:er|re)s)\s*-\s*([0-9.,]+)", ltext)
                    if m:
                        num = m.group(1).replace(",", "")
                        lm = float(num)
                        spec = pending_specs[-1] if pending_specs else "Edgeband"
                        band_items.append(BandItem(spec=spec, lm=lm))
                except Exception:
                    pass
            i += 1

    # Parse hardware totals: handle multiple blocks
    for anchor in hw_anchors:
        i = anchor + 1
        while i < len(rows):
            row = rows[i]
            cells = _cell_texts(row)
            text = " ".join(cells)
            ltext = text.lower()
            if not text:
                i += 1
                continue
            # stop if next section appears
            if any(h in ltext for h in [
                "sheet stock totals",
                "edgeband totals",
                "work order summary",
                "project name:",
                "batch:",
                "buyout totals",
            ]):
                break
            if "qty -" in ltext:
                try:
                    qty = int(ltext.split("qty -", 1)[1].split()[0].replace(",", ""))
                    # The item description: prefer the longest cell that is not the qty token
                    desc_candidates = [c for c in cells if "qty -" not in c.lower()]
                    desc = max(desc_candidates, key=len, default="").strip()
                    if desc:
                        hw_items.append(HardwareItem(description=desc, qty=qty))
                except Exception:
                    pass
            i += 1

    return {
        "sheets": [s.__dict__ for s in sheet_items],
        "edgeband": [b.__dict__ for b in band_items],
        "hardware": [h.__dict__ for h in hw_items],
    }
